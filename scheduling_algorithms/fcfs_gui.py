''' 
Python 3.13.5
Code: Paul Mendoza
'''

import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import os

def fcfs(processes: List[Dict]):
    indexed = [(i, p["pid"], p["AT"], p["BT"]) for i, p in enumerate(processes)]
    indexed.sort(key=lambda x: (x[2], x[0]))

    time = 0
    rows = []
    timeline = []
    steps = []

    steps.append("Step 1: Sort processes by Arrival Time (AT)")
    step_detail = "Sorted order: " + ", ".join([f"{pid} (AT={at})" for _, pid, at, _ in indexed])
    steps.append(step_detail)
    steps.append("")

    for idx, (_, pid, at, bt) in enumerate(indexed):
        steps.append(f"Step {idx+2}: Processing {pid} (AT={at}, BT={bt})")
        
        if time < at:
            idle_time = at - time
            steps.append(f"  - CPU is idle for {idle_time} units (from {time} to {at})")
            timeline.append(("IDLE", time, at))
            time = at

        st = time
        ct = st + bt
        wt = st - at
        tat = ct - at
        rt = wt

        steps.append(f"  - Start Time (ST): {st}")
        steps.append(f"  - Completion Time (CT): {st} + {bt} = {ct}")
        steps.append(f"  - Waiting Time (WT): {st} - {at} = {wt}")
        steps.append(f"  - Turnaround Time (TAT): {ct} - {at} = {tat}")
        steps.append(f"  - Response Time (RT): {wt} (same as WT in FCFS)")
        steps.append("")

        rows.append({
            "PID": pid, "AT": at, "BT": bt,
            "ST": st, "CT": ct, "WT": wt, "TAT": tat, "RT": rt
        })
        timeline.append((pid, st, ct))
        time = ct

    n = len(rows)
    tbt = sum(r["BT"] for r in rows)
    tft = timeline[-1][2] if timeline else 0
    busy_time = sum(e - s for (lab, s, e) in timeline if lab != "IDLE")

    avg_wt = sum(r["WT"] for r in rows) / n if n > 0 else 0
    avg_tat = sum(r["TAT"] for r in rows) / n if n > 0 else 0
    cpu_util = (busy_time / tft * 100) if tft > 0 else 0
    throughput = (n / tft) if tft > 0 else 0

    steps.append("Step Summary:")
    steps.append(f"  - Average Waiting Time (AWT): {avg_wt:.2f}")
    steps.append(f"  - Average Turnaround Time (ATAT): {avg_tat:.2f}")
    steps.append(f"  - CPU Utilization: {cpu_util:.2f}%")
    steps.append(f"  - Throughput: {throughput:.4f} processes/unit time")

    return rows, timeline, avg_wt, avg_tat, tbt, tft, cpu_util, throughput, steps

def export_to_excel():
    os.makedirs("output",exist_ok=True)
    if not processes:
        messagebox.showwarning("No Data", "No processes to export!")
        return
        
    rows, timeline, avg_wt, avg_tat, tbt, tft, cpu_util, throughput, steps = fcfs(processes)
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "FCFS Results"
        
        headers = ["PID", "AT", "BT", "ST", "CT", "WT", "TAT", "RT"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        for row_idx, row in enumerate(rows, 2):
            ws.cell(row=row_idx, column=1, value=row["PID"])
            ws.cell(row=row_idx, column=2, value=row["AT"])
            ws.cell(row=row_idx, column=3, value=row["BT"])
            ws.cell(row=row_idx, column=4, value=row["ST"])
            ws.cell(row=row_idx, column=5, value=row["CT"])
            ws.cell(row=row_idx, column=6, value=row["WT"])
            ws.cell(row=row_idx, column=7, value=row["TAT"])
            ws.cell(row=row_idx, column=8, value=row["RT"])
        
        ws.cell(row=len(rows)+3, column=1, value="Summary").font = Font(bold=True)
        ws.cell(row=len(rows)+4, column=1, value="Average Waiting Time (AWT)")
        ws.cell(row=len(rows)+4, column=2, value=avg_wt)
        ws.cell(row=len(rows)+5, column=1, value="Average Turnaround Time (ATAT)")
        ws.cell(row=len(rows)+5, column=2, value=avg_tat)
        ws.cell(row=len(rows)+6, column=1, value="Total Burst Time (TBT)")
        ws.cell(row=len(rows)+6, column=2, value=tbt)
        ws.cell(row=len(rows)+7, column=1, value="Total Finish Time (TFT)")
        ws.cell(row=len(rows)+7, column=2, value=tft)
        ws.cell(row=len(rows)+8, column=1, value="CPU Utilization (%)")
        ws.cell(row=len(rows)+8, column=2, value=cpu_util)
        ws.cell(row=len(rows)+9, column=1, value="Throughput")
        ws.cell(row=len(rows)+9, column=2, value=throughput)
        
        filename = "output/fcfs_results.xlsx"
        wb.save(filename)
        messagebox.showinfo("Export Successful", f"Results exported to {filename}")
        
    except Exception as e:
        messagebox.showerror("Export Error", f"Failed to export to Excel: {str(e)}")

def add_process():
    try:
        at = int(at_var.get())
        bt = int(bt_var.get())
        if at < 0 or bt <= 0:
            raise ValueError
    except ValueError:
        messagebox.showerror("Invalid Input", "AT must be >=0 and BT > 0")
        return

    pid = f"P{len(processes) + 1}"
    processes.append({"pid": pid, "AT": at, "BT": bt})
    tree.insert("", "end", values=(pid, at, bt))

    clear_fields()

def clear_fields():
    at_var.set("")
    bt_var.set("")
    
def clear_all():
    global processes
    processes.clear()
    for item in tree.get_children():
        tree.delete(item)
    results_canvas.delete("all")
    step_text.delete(1.0, tk.END)
    export_btn.config(state=tk.DISABLED)

def run_fcfs():
    if not processes:
        messagebox.showwarning("No Data", "Please add processes first!")
        return

    rows, timeline, avg_wt, avg_tat, tbt, tft, cpu_util, throughput, steps = fcfs(processes)

    step_text.delete(1.0, tk.END)
    for step in steps:
        step_text.insert(tk.END, step + "\n")

    results_canvas.delete("all")
    
    results_frame = tk.Frame(results_canvas, bg="white")
    results_canvas.create_window((0, 0), window=results_frame, anchor="nw")
    
    tk.Label(results_frame, text="FCFS Scheduling Results", font=("Arial", 14, "bold"), bg="white").grid(row=0, column=0, columnspan=8, pady=10)
    
    headers = ["PID", "AT", "BT", "ST", "CT", "WT", "TAT", "RT"]
    for col, header in enumerate(headers):
        tk.Label(results_frame, text=header, font=("Arial", 10, "bold"), bg="lightgray", width=8).grid(row=1, column=col, padx=2, pady=2)
    
    for row_idx, row in enumerate(rows):
        tk.Label(results_frame, text=row["PID"], bg="white", width=8).grid(row=row_idx+2, column=0, padx=2, pady=1)
        tk.Label(results_frame, text=row["AT"], bg="white", width=8).grid(row=row_idx+2, column=1, padx=2, pady=1)
        tk.Label(results_frame, text=row["BT"], bg="white", width=8).grid(row=row_idx+2, column=2, padx=2, pady=1)
        tk.Label(results_frame, text=row["ST"], bg="white", width=8).grid(row=row_idx+2, column=3, padx=2, pady=1)
        tk.Label(results_frame, text=row["CT"], bg="white", width=8).grid(row=row_idx+2, column=4, padx=2, pady=1)
        tk.Label(results_frame, text=row["WT"], bg="white", width=8).grid(row=row_idx+2, column=5, padx=2, pady=1)
        tk.Label(results_frame, text=row["TAT"], bg="white", width=8).grid(row=row_idx+2, column=6, padx=2, pady=1)
        tk.Label(results_frame, text=row["RT"], bg="white", width=8).grid(row=row_idx+2, column=7, padx=2, pady=1)
    
    start_row = len(rows) + 3
    tk.Label(results_frame, text="Summary", font=("Arial", 12, "bold"), bg="white").grid(row=start_row, column=0, columnspan=8, pady=10)
    
    tk.Label(results_frame, text="Average Waiting Time (AWT):", font=("Arial", 10), bg="white").grid(row=start_row+1, column=0, columnspan=4, sticky="e")
    tk.Label(results_frame, text=f"{avg_wt:.2f}", font=("Arial", 10), bg="white").grid(row=start_row+1, column=4, columnspan=4, sticky="w")
    
    tk.Label(results_frame, text="Average Turnaround Time (ATAT):", font=("Arial", 10), bg="white").grid(row=start_row+2, column=0, columnspan=4, sticky="e")
    tk.Label(results_frame, text=f"{avg_tat:.2f}", font=("Arial", 10), bg="white").grid(row=start_row+2, column=4, columnspan=4, sticky="w")
    
    tk.Label(results_frame, text="CPU Utilization:", font=("Arial", 10), bg="white").grid(row=start_row+3, column=0, columnspan=4, sticky="e")
    tk.Label(results_frame, text=f"{cpu_util:.2f}%", font=("Arial", 10), bg="white").grid(row=start_row+3, column=4, columnspan=4, sticky="w")
    
    tk.Label(results_frame, text="Throughput:", font=("Arial", 10), bg="white").grid(row=start_row+4, column=0, columnspan=4, sticky="e")
    tk.Label(results_frame, text=f"{throughput:.4f}", font=("Arial", 10), bg="white").grid(row=start_row+4, column=4, columnspan=4, sticky="w")
    
    results_frame.update_idletasks()
    results_canvas.configure(scrollregion=results_canvas.bbox("all"))
    
    # draw_gantt_chart(timeline)
    
    export_btn.config(state=tk.NORMAL)

def open_gantt_window(timeline):
    if not timeline:
        messagebox.showwarning("No Data", "Run FCFS first to generate the Gantt chart!")
        return

    gantt_win = tk.Toplevel(root)
    gantt_win.title("Gantt Chart")
    gantt_win.geometry("800x200")

    gantt_frame = ttk.Frame(gantt_win)
    gantt_frame.pack(fill="both", expand=True, padx=10, pady=10)

    gantt_canvas = tk.Canvas(gantt_frame, height=100, bg="white")
    h_scrollbar = ttk.Scrollbar(gantt_frame, orient="horizontal", command=gantt_canvas.xview)
    gantt_canvas.configure(xscrollcommand=h_scrollbar.set)

    gantt_canvas.pack(side="top", fill="x", expand=True)
    h_scrollbar.pack(side="bottom", fill="x")

    total_time = timeline[-1][2]
    x = 10
    for label, s, e in timeline:
        w = max(30, (e - s) * 30)
        color = "lightblue" if label != "IDLE" else "lightgray"
        gantt_canvas.create_rectangle(x, 20, x + w, 60, fill=color, outline="black")
        gantt_canvas.create_text(x + w / 2, 40, text=label)
        gantt_canvas.create_text(x, 70, text=str(s), anchor="nw")
        x += w

    gantt_canvas.create_text(x, 70, text=str(total_time), anchor="nw")
    gantt_canvas.configure(scrollregion=(0, 0, x + 50, 100))


# def draw_gantt_chart(timeline):
#     gantt_canvas.delete("all")
#     if not timeline:
#         return
        
#     total_time = timeline[-1][2]
    
#     x = 10
#     for label, s, e in timeline:
#         w = max(30, (e - s) * 30)
#         color = "lightblue" if label != "IDLE" else "lightgray"
#         gantt_canvas.create_rectangle(x, 20, x + w, 60, fill=color, outline="black")
#         gantt_canvas.create_text(x + w / 2, 40, text=label)
#         gantt_canvas.create_text(x, 70, text=str(s), anchor="nw")
#         x += w
        
#     gantt_canvas.create_text(x, 70, text=str(total_time), anchor="nw")
    
#     gantt_canvas.configure(scrollregion=(0, 0, x + 50, 100))
    


root = tk.Tk()
style = ttk.Style(root)
style.theme_use("clam")
root.title("FCFS Scheduling - GUI By Paul Mendoza")
# root.geometry("1000x800")
# root.resizable(False, False)
# root.attributes("-zoomed", True)

processes = []


frame_input = ttk.LabelFrame(root, text="Process Input")
frame_input.pack(fill="x", padx=10, pady=5)

tree_frame = ttk.Frame(frame_input)
tree_frame.pack(fill="both", expand=True)

tree = ttk.Treeview(tree_frame, columns=("PID", "AT", "BT"), show="headings", height=5)
tree.heading("PID", text="Process ID")
tree.heading("AT", text="Arrival Time (AT)")
tree.heading("BT", text="Burst Time (BT)")
tree.column("PID", width=100, anchor="center")
tree.column("AT", width=120, anchor="center")
tree.column("BT", width=120, anchor="center")
tree.pack(side="left", fill="both", expand=True)

scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
tree.configure(yscroll=scrollbar.set)
scrollbar.pack(side="right", fill="y")

at_var = tk.StringVar()
bt_var = tk.StringVar()

entry_frame = ttk.Frame(root)
entry_frame.pack(pady=5)
ttk.Label(entry_frame, text="Arrival Time:").grid(row=0, column=0, padx=5)
ttk.Entry(entry_frame, textvariable=at_var, width=10).grid(row=0, column=1, padx=5)
ttk.Label(entry_frame, text="Burst Time:").grid(row=0, column=2, padx=5)
ttk.Entry(entry_frame, textvariable=bt_var, width=10).grid(row=0, column=3, padx=5)
ttk.Button(entry_frame, text="Add Process", command=add_process).grid(row=0, column=4, padx=5)
ttk.Button(entry_frame, text="Clear All", command=clear_all).grid(row=0, column=5, padx=5)

button_frame = ttk.Frame(root)
button_frame.pack(pady=10)
ttk.Button(button_frame, text="Run FCFS", command=run_fcfs).grid(row=0, column=0, padx=5)
ttk.Button(button_frame, text="Show Gantt Chart", command=lambda: open_gantt_window(fcfs(processes)[1])).grid(row=0, column=2, padx=5)
export_btn = ttk.Button(button_frame, text="Export to Excel", command=export_to_excel, state=tk.DISABLED)
export_btn.grid(row=0, column=1, padx=5)

notebook = ttk.Notebook(root)
notebook.pack(fill="both", expand=True, padx=10, pady=5)

step_frame = ttk.Frame(notebook)
notebook.add(step_frame, text="Step-by-Step Solution")

step_text = scrolledtext.ScrolledText(step_frame, wrap=tk.WORD, width=80, height=15)
step_text.pack(fill="both", expand=True, padx=5, pady=5)

results_frame = ttk.Frame(notebook)
notebook.add(results_frame, text="Results")

results_canvas_frame = ttk.Frame(results_frame)
results_canvas_frame.pack(fill="both", expand=True)

results_canvas = tk.Canvas(results_canvas_frame, bg="white")
v_scrollbar = ttk.Scrollbar(results_canvas_frame, orient="vertical", command=results_canvas.yview)
h_scrollbar = ttk.Scrollbar(results_canvas_frame, orient="horizontal", command=results_canvas.xview)
results_canvas.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)

results_canvas.grid(row=0, column=0, sticky="nsew")
v_scrollbar.grid(row=0, column=1, sticky="ns")
h_scrollbar.grid(row=1, column=0, sticky="ew")

results_canvas_frame.grid_rowconfigure(0, weight=1)
results_canvas_frame.grid_columnconfigure(0, weight=1)

# gantt_frame = ttk.LabelFrame(root, text="Gantt Chart")
# gantt_frame.pack(fill="x", padx=10, pady=10)

# gantt_canvas_frame = ttk.Frame(gantt_frame)
# gantt_canvas_frame.pack(fill="x", expand=True)

# gantt_canvas = tk.Canvas(gantt_canvas_frame, height=100, bg="white")
# gantt_scrollbar = ttk.Scrollbar(gantt_canvas_frame, orient="horizontal", command=gantt_canvas.xview)
# gantt_canvas.configure(xscrollcommand=gantt_scrollbar.set)

# gantt_canvas.pack(side="top", fill="x", expand=True)
# gantt_scrollbar.pack(side="bottom", fill="x")

root.mainloop()